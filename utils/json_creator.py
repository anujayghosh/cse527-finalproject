import openpyxl
import json

def read_points_sheet(sheet):
    points = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        x, y = row[1], row[2]
        points.append([x, y])
    return points

def calculate_rectangle_coordinates(row):
    x1 = row[1]
    y1 = row[2]
    width = row[3]
    height = row[4]
    x2 = x1 
    y2 = y1 + height
    x3 = x1 + width
    y3 = y1 + height
    x4 = x1 + width
    y4 = y1

    return [[x1, y1], [x2, y2], [x3, y3], [x4,y4]]

def read_rectangles_sheet(sheet):
    rectangles = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        coordinates = calculate_rectangle_coordinates(row)
        rectangles.append(coordinates)
    return rectangles

def get_image_width_height(sheet):
    
    column_index_width = 7
    column_index_height = 8
    
    # Assuming you want to extract the height from the second row
    row_index = 2
    
    # Access the cell containing the height value
    height_value = sheet.cell(row=row_index, column=column_index_height).value
    width_value = sheet.cell(row=row_index, column=column_index_width).value
    
    return height_value, width_value

def excel_to_json(file_path_list):
    all_data = {}  # Initialize a dictionary to store data for all files
    for file_path in file_path_list:
        workbook = openpyxl.load_workbook(file_path)
        
        points_sheet = workbook['Sheet1']
        points = read_points_sheet(points_sheet)
        
        rectangles_sheet = workbook['Sheet2']
        rectangles = read_rectangles_sheet(rectangles_sheet)

        height, width = get_image_width_height(rectangles_sheet)

        file_name_image = file_path[:-5]+".jpg"
        
        # Add data for the current file to the dictionary
        all_data[file_name_image] = {
            "H": height,
            "W": width,
            "box_examples_coordinates": rectangles,
            "box_examples_path": [
                f"/nfs/bigneuron/viresh/FSC_NewDataOnly/box_examples/{file_path[:4]}_0.jpg",
                f"/nfs/bigneuron/viresh/FSC_NewDataOnly/box_examples/{file_path[:4]}_1.jpg"
            ],
            "density_path": f"/nfs/bigneuron/viresh/FSC_NewDataOnly/gt_density_map_adaptive_384_VarV2/{file_path[:4]}.npy",
            "density_path_fixed": f"/nfs/bigneuron/viresh/FSC_NewDataOnly/gt_density_map_fixed/{file_path[:4]}.npy",
            "img_path": f"/nfs/bigneuron/viresh/FSC_NewDataOnly/images_384_VarV2/{file_name_image}",
            "points": points,
            "r": [15,15],
            "ratio_h": 1.0,
            "ratio_w": 1.6015625
        }
    
    return json.dumps(all_data, ensure_ascii=False, indent=4)

if __name__ == "__main__":
    file_name_list = ['8006','8008','8011','8012','8015','8016','8017','8018','8019','8020','8021','8022','8023','8024']
    file_path_list = [x+".xlsx" for x in file_name_list ] # Replace with the list of file paths
    output_file = "output.json"

    with open(output_file, "w") as f:  # Use "w" for text mode
        json_output = excel_to_json(file_path_list)
        f.write(json_output)

    print("JSON data saved to", output_file)







print("JSON data saved to", output_file)